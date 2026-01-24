import os
import sys
import pandas as pd

CSV_NAME = "booking_result.csv"
XLSX_NAME = "booking_result.xlsx"


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    csv_path = os.path.join(here, CSV_NAME)
    xlsx_path = os.path.join(here, XLSX_NAME)

    if not os.path.exists(csv_path):
        print(f"⛔ ไม่พบไฟล์ {CSV_NAME} ในโฟลเดอร์: {here}")
        sys.exit(1)

    # โหลดข้อมูลจาก CSV
    try:
        df = pd.read_csv(csv_path, encoding="utf-8-sig")
    except UnicodeDecodeError:
        # เผื่อกรณี encoding แตกต่าง
        df = pd.read_csv(csv_path, encoding="utf-8")

    # จัดชนิดข้อมูลคอลัมน์ 'วันที่' ให้เป็นตัวเลข (ถ้าเป็นไปได้)
    if "วันที่" in df.columns:
        df["วันที่"] = pd.to_numeric(df["วันที่"], errors="coerce")

    # เลือก engine สำหรับเขียน Excel
    engine = None
    try:
        import xlsxwriter  # noqa: F401
        engine = "xlsxwriter"
    except Exception:
        try:
            import openpyxl  # noqa: F401
            engine = "openpyxl"
        except Exception:
            print("⛔ ไม่พบไลบรารีสำหรับเขียน Excel (openpyxl/xlsxwriter)")
            print("โปรดติดตั้งด้วยคำสั่ง: python -m pip install openpyxl xlsxwriter")
            sys.exit(1)

    # เขียนเป็นไฟล์ Excel พร้อมจัดรูปแบบพื้นฐาน
    with pd.ExcelWriter(xlsx_path, engine=engine) as writer:
        sheet_name = "Bookings"
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        # จัดรูปแบบผ่าน engine หากรองรับ
        try:
            worksheet = writer.sheets[sheet_name]

            # Freeze แถวหัวตาราง
            if engine == "xlsxwriter":
                worksheet.freeze_panes(1, 0)
            elif engine == "openpyxl":
                worksheet.freeze_panes = "A2"

            # ตั้งความกว้างคอลัมน์แบบเหมาะสม
            col_widths = {
                "เว็บไซต์": 22,
                "ชื่อบ้าน": 30,
                "รหัส": 14,
                "เดือน": 18,
                "วันที่": 10,
                "สถานะ": 12,
            }

            # ทำแผนที่ column index -> width
            for idx, col in enumerate(df.columns):
                width = col_widths.get(col, max(10, min(40, len(col) + 2)))
                # แปลง index เป็นรูปแบบ A, B, C เฉพาะ openpyxl
                if engine == "xlsxwriter":
                    worksheet.set_column(idx, idx, width)
                else:
                    from openpyxl.utils import get_column_letter
                    worksheet.column_dimensions[get_column_letter(idx + 1)].width = width

            # ใส่ AutoFilter ให้หัวตาราง
            if engine == "xlsxwriter":
                worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)
            else:
                # openpyxl: เปิด AutoFilter โดยกำหนด ref ครอบคลุมตาราง
                from openpyxl.utils import get_column_letter
                last_col = get_column_letter(len(df.columns))
                last_row = len(df) + 1  # รวม header
                worksheet.auto_filter.ref = f"A1:{last_col}{last_row}"

        except Exception as e:
            # ไม่ให้ล้ม หากจัดรูปแบบไม่ได้
            print(f"⚠️ จัดรูปแบบเวิร์กชีตไม่สำเร็จ: {e}")

    print(f"✅ สร้างไฟล์ Excel สำเร็จ → {XLSX_NAME}")


if __name__ == "__main__":
    main()
