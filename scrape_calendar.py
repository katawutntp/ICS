from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
import re
import os
from datetime import datetime
from dateutil.relativedelta import relativedelta
from urllib.parse import urlparse

# ===== CONFIG =====
MONTH_TO_SCRAPE = 5   # 👈 จำนวนเดือนที่ต้องการ
MAX_HOUSES = 0        # 👈 จำนวนบ้านที่ต้องการดึง (0 = ทั้งหมด)

# รายการ URL ที่ต้องการ scrape (รองรับหลายเว็บ) — ใช้เป็น fallback
URLS = [
     "https://www.devillegroups.com/allcalendar/?s=1758",  # รูปแบบที่ 1: Deville Groups
    "https://poolvillacity.co.th/CITY-743",               # รูปแบบที่ 2: Pool Villa City
     "https://www.pattayapartypoolvilla.com/v/2246",       # รูปแบบที่ 3: Pattaya Party Pool Villa
]
# ==================


def load_urls_from_webpath():
    """อ่านรายการ URL จากไฟล์ webpath ที่อยู่ในโฟลเดอร์เดียวกัน
    รูปแบบแต่ละบรรทัด: ชื่อ\tURL หรือ ชื่อ URL (คั่นด้วยช่องว่าง)
    ข้ามบรรทัดว่างและบรรทัดคอมเมนต์ (# ...)
    """
    here = os.path.dirname(os.path.abspath(__file__))
    webpath = os.path.join(here, "webpath")

    urls = []
    if not os.path.exists(webpath):
        return urls

    with open(webpath, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            # หา URL ด้วย regex (http:// หรือ https://)
            url_match = re.search(r'https?://[^\s]+', line)
            if url_match:
                urls.append(url_match.group(0))
    return urls


class CalendarScraper:
    """Base class สำหรับ scraping ปฏิทิน"""
    
    # Map ชื่อเดือนภาษาไทย -> เลขเดือน
    THAI_MONTH_MAP = {
        'มกราคม': 1, 'กุมภาพันธ์': 2, 'มีนาคม': 3, 'เมษายน': 4,
        'พฤษภาคม': 5, 'มิถุนายน': 6, 'กรกฎาคม': 7, 'สิงหาคม': 8,
        'กันยายน': 9, 'ตุลาคม': 10, 'พฤศจิกายน': 11, 'ธันวาคม': 12
    }
    
    def __init__(self, driver):
        self.driver = driver
        self.results = []
        self.today = datetime.now().date()  # วันที่ปัจจุบัน
    
    def filter_past_dates(self, results):
        """
        กรองวันที่ก่อนวันปัจจุบันออก (ใช้ได้กับทุกเว็บไซต์)
        
        รองรับรูปแบบเดือน:
        - "มกราคม 2569" (พ.ศ.)
        - "2026-01" (ค.ศ.)
        """
        filtered = []
        
        for row in results:
            try:
                month_str = row.get('เดือน', '')
                day = int(row.get('วันที่', 0))
                
                # แปลงเดือนเป็น year, month
                year, month = self._parse_month_string(month_str)
                
                if year and month and day:
                    # สร้าง date object
                    from datetime import date
                    row_date = date(year, month, day)
                    
                    # เก็บเฉพาะวันที่ >= วันปัจจุบัน
                    if row_date >= self.today:
                        filtered.append(row)
            except Exception as e:
                # ถ้า parse ไม่ได้ ให้เก็บไว้ก่อน
                filtered.append(row)
        
        return filtered
    
    def _parse_month_string(self, month_str):
        """
        แปลง string เดือนเป็น (year, month)
        
        รองรับ:
        - "มกราคม 2569" -> (2026, 1)
        - "2026-01" -> (2026, 1)
        """
        try:
            # รูปแบบ "มกราคม 2569"
            for thai_month, month_num in self.THAI_MONTH_MAP.items():
                if thai_month in month_str:
                    # หาปี พ.ศ.
                    year_match = re.search(r'(25\d{2}|26\d{2}|27\d{2})', month_str)
                    if year_match:
                        thai_year = int(year_match.group(1))
                        # แปลง พ.ศ. เป็น ค.ศ.
                        year = thai_year - 543
                        return (year, month_num)
            
            # รูปแบบ "2026-01"
            match = re.match(r'(\d{4})-(\d{2})', month_str)
            if match:
                year = int(match.group(1))
                month = int(match.group(2))
                return (year, month)
                
        except:
            pass
        
        return (None, None)
    
    def detect_site_type(self, url):
        """ตรวจจับประเภทเว็บไซต์จาก URL"""
        domain = urlparse(url).netloc.lower()
        
        if 'devillegroups.com' in domain:
            return 'deville'
        elif 'poolvillacity.co.th' in domain:
            return 'poolvillacity'
        elif 'pattayapartypoolvilla.com' in domain:
            return 'pattayaparty'
        else:
            return 'unknown'
    
    def scrape(self, url):
        """เลือก scraper ตามประเภทเว็บไซต์"""
        site_type = self.detect_site_type(url)
        
        print(f"\n{'='*60}")
        print(f"🌐 URL: {url}")
        print(f"📌 ประเภท: {site_type}")
        print(f"{'='*60}")
        
        if site_type == 'deville':
            return self.scrape_deville(url)
        elif site_type == 'poolvillacity':
            return self.scrape_poolvillacity(url)
        elif site_type == 'pattayaparty':
            return self.scrape_pattayaparty(url)
        else:
            print(f"❌ ไม่รู้จักประเภทเว็บไซต์: {url}")
            return []


    # ========================================================
    # รูปแบบที่ 1: Deville Groups (หลายบ้านในหน้าเดียว + iframe)
    # ========================================================
    def scrape_deville(self, url):
        """Scrape ปฏิทินจาก devillegroups.com"""
        print("🔄 กำลังโหลดหน้าหลัก Deville Groups...")
        
        BASE_IFRAME_URL = "https://www.devillegroups.com/allcalendar/cld.php"
        
        self.driver.get(url)
        time.sleep(8)
        
        results = []
        html = self.driver.page_source
        
        # หา pattern: <h6>(DV-xxxx)<br>ชื่อบ้าน</h6>...<iframe src="cld.php?hId=xxxx"
        pattern = r'<h6>\(DV-(\d+)\)<br>([^<]+)</h6>.*?src="cld\.php\?hId=(\d+)"'
        matches = re.findall(pattern, html, re.DOTALL)
        
        houses = []
        seen_ids = set()
        
        for dv_id, name, h_id in matches:
            if h_id in seen_ids:
                continue
            seen_ids.add(h_id)
            
            house_name = name.strip()
            houses.append({
                'id': h_id,
                'name': house_name,
                'dv_code': f'DV-{dv_id}'
            })
            print(f"  🏠 พบบ้าน: {house_name} (DV-{dv_id}, hId={h_id})")
        
        print(f"\n📊 พบบ้านทั้งหมด: {len(houses)} หลัง")
        
        if not houses:
            print("❌ ไม่พบข้อมูลบ้าน")
            return results
        
        # จำกัดจำนวนบ้าน
        if MAX_HOUSES > 0:
            houses = houses[:MAX_HOUSES]
            print(f"🔧 จำกัดดึงแค่ {MAX_HOUSES} หลังแรก")
        
        # วนดึงข้อมูลแต่ละบ้าน
        start_date = datetime.now()
        total_houses = len(houses)
        
        for house_idx, house in enumerate(houses, 1):
            h_id = house['id']
            house_name = house['name']
            dv_code = house['dv_code']
            
            print(f"\n{'='*50}")
            print(f"🏠 [{house_idx}/{total_houses}] กำลังดึง: {house_name} ({dv_code})")
            print(f"{'='*50}")
            
            for i in range(MONTH_TO_SCRAPE):
                try:
                    target_date = start_date + relativedelta(months=i)
                    ym = target_date.strftime("%Y-%m")
                    calendar_url = f"{BASE_IFRAME_URL}?ym={ym}&hId={h_id}"
                    
                    self.driver.get(calendar_url)
                    time.sleep(2)
                    
                    wait = WebDriverWait(self.driver, 10)
                    
                    # อ่านชื่อเดือน
                    try:
                        month_el = wait.until(
                            EC.presence_of_element_located(
                                (By.XPATH, "//th[contains(text(),'256') or contains(text(),'257')]")
                            )
                        )
                        month_text = month_el.text.strip()
                        for line in month_text.split("\n"):
                            if "256" in line or "257" in line:
                                month_text = line.strip()
                                break
                    except:
                        month_text = ym
                    
                    # ดึงวันที่ติดจอง (สีแดง = booking)
                    booked_cells = self.driver.find_elements(
                        By.XPATH,
                        "//td[contains(@class,'booking') or contains(@style,'red')]"
                    )
                    
                    booked_count = 0
                    booked_days = []
                    for cell in booked_cells:
                        day = cell.text.strip()
                        if day.isdigit():
                            booked_days.append(int(day))
                            results.append({
                                "เว็บไซต์": "Deville Groups",
                                "ชื่อบ้าน": house_name,
                                "รหัส": dv_code,
                                "เดือน": month_text,
                                "วันที่": int(day),
                                "สถานะ": "ติดจอง"
                            })
                            booked_count += 1
                    
                    if booked_days:
                        days_str = ', '.join(map(str, sorted(booked_days)))
                        print(f"  📅 {month_text}: {booked_count} วัน → [{days_str}]")
                    else:
                        print(f"  📅 {month_text}: ว่าง ✓")
                        
                except Exception as e:
                    print(f"  ⛔ Error ({ym}): {e}")
        
        return results


    # ========================================================
    # รูปแบบที่ 2: Pool Villa City (ปฏิทิน FullCalendar + navigation)
    # ========================================================
    def scrape_poolvillacity(self, url):
        """
        Scrape ปฏิทินจาก poolvillacity.co.th
        
        ✅ รองรับ:
        - ใช้ FullCalendar library
        - ปฏิทินแสดงทีละหลายเดือน (ต้องกด Next เพื่อดูเดือนถัดไป)
        - วันที่ติดจอง: มี fc-bg-event + background-color: rgb(248, 229, 231) + สีแดง
        - วันที่เทศกาล: สีเหลือง
        - วันที่อยู่ใน data-date attribute
        """
        print("🔄 กำลังโหลดหน้า Pool Villa City...")
        
        results = []
        
        # ดึงรหัสบ้านจาก URL (เช่น CITY-743)
        match = re.search(r'(CITY-\d+)', url)
        if match:
            house_code = match.group(1)
        else:
            house_code = "Unknown"
        
        self.driver.get(url)
        time.sleep(8)  # รอ JavaScript โหลด
        
        try:
            wait = WebDriverWait(self.driver, 15)
            
            # ดึงชื่อบ้าน
            try:
                title_el = wait.until(
                    EC.presence_of_element_located((By.TAG_NAME, "h1"))
                )
                house_name = title_el.text.strip()
                if not house_name:
                    house_name = house_code
            except:
                house_name = house_code
            
            print(f"  🏠 บ้าน: {house_name} ({house_code})")
            
            # รอให้ FullCalendar โหลด
            try:
                wait.until(
                    EC.presence_of_element_located((By.CLASS_NAME, "fc-daygrid-day"))
                )
            except:
                print("  ⚠️ ไม่พบ FullCalendar - รอเพิ่ม...")
                time.sleep(5)
            
            booked_dates = set()  # ใช้ set เพื่อไม่ซ้ำ
            
            month_map = {
                '01': 'มกราคม', '02': 'กุมภาพันธ์', '03': 'มีนาคม',
                '04': 'เมษายน', '05': 'พฤษภาคม', '06': 'มิถุนายน',
                '07': 'กรกฎาคม', '08': 'สิงหาคม', '09': 'กันยายน',
                '10': 'ตุลาคม', '11': 'พฤศจิกายน', '12': 'ธันวาคม'
            }
            
            current_year = datetime.now().year
            current_month = datetime.now().month
            
            # วนกดปุ่ม Next เพื่อดึงข้อมูลหลายรอบ
            # FullCalendar อาจแสดงหลายเดือนในหน้าเดียว เราจะกด Next หลายครั้ง
            for round_num in range(MONTH_TO_SCRAPE):
                # หา td ที่มี data-date และมี fc-bg-event ด้านใน (วันที่ติดจอง)
                booked_cells = self.driver.find_elements(
                    By.XPATH,
                    "//td[contains(@class,'fc-daygrid-day')]//div[contains(@class,'fc-bg-event') and contains(@style,'rgb(248, 229, 231)')]"
                )
                
                # ถ้าไม่พบ ลองหาจาก parent td
                if not booked_cells:
                    all_day_cells = self.driver.find_elements(
                        By.XPATH,
                        "//td[contains(@class,'fc-daygrid-day') and .//div[contains(@class,'fc-bg-event')]]"
                    )
                    booked_cells = all_day_cells
                
                for cell in booked_cells:
                    try:
                        data_date = cell.get_attribute("data-date")
                        if not data_date:
                            parent_td = cell.find_element(By.XPATH, "./ancestor::td[@data-date]")
                            data_date = parent_td.get_attribute("data-date")
                        
                        if data_date:
                            booked_dates.add(data_date)
                    except:
                        pass
                
                # กดปุ่ม Next เพื่อไปเดือนถัดไป (ยกเว้นรอบสุดท้าย)
                if round_num < MONTH_TO_SCRAPE - 1:
                    try:
                        next_btn = self.driver.find_element(
                            By.XPATH,
                            "//button[contains(@class,'fc-next-button') or contains(@aria-label,'next') or contains(@title,'Next')]"
                        )
                        next_btn.click()
                        time.sleep(2)
                    except:
                        # ถ้าไม่พบปุ่มก็หยุด
                        break
            
            # จัดกลุ่มตามเดือน
            by_month = {}
            for date_str in sorted(booked_dates):
                parts = date_str.split('-')
                if len(parts) == 3:
                    year, month, day = parts
                    year_int = int(year)
                    month_int = int(month)
                    
                    # ข้ามเดือนที่ผ่านมาแล้ว
                    if year_int < current_year or (year_int == current_year and month_int < current_month):
                        continue
                    
                    # จำกัดแค่ MONTH_TO_SCRAPE เดือน
                    months_diff = (year_int - current_year) * 12 + (month_int - current_month)
                    if months_diff >= MONTH_TO_SCRAPE:
                        continue
                    
                    # แปลงปี ค.ศ. เป็น พ.ศ.
                    thai_year = int(year) + 543
                    month_name = month_map.get(month, month)
                    month_key = f"{month_name} {thai_year}"
                    
                    if month_key not in by_month:
                        by_month[month_key] = []
                    by_month[month_key].append(int(day))
                    
                    results.append({
                        "เว็บไซต์": "Pool Villa City",
                        "ชื่อบ้าน": house_name,
                        "รหัส": house_code,
                        "เดือน": month_key,
                        "วันที่": int(day),
                        "สถานะ": "ติดจอง"
                    })
            
            # แสดงผล
            if by_month:
                for month_key, days in sorted(by_month.items()):
                    days_str = ', '.join(map(str, sorted(days)))
                    print(f"  📅 {month_key}: {len(days)} วัน → [{days_str}]")
            else:
                print("  📅 ไม่พบวันติดจอง (ว่างทั้งหมด หรืออาจต้องปรับ selector)")
                
        except Exception as e:
            print(f"  ⛔ Error: {e}")
            import traceback
            traceback.print_exc()
        
        return results


    # ========================================================
    # รูปแบบที่ 3: Pattaya Party Pool Villa (ปฏิทินเดือนเดียว + navigation)
    # ========================================================
    def scrape_pattayaparty(self, url):
        """
        Scrape ปฏิทินจาก pattayapartypoolvilla.com
        
        ✅ รองรับ:
        - ปฏิทินแสดงทีละเดือน
        - มีปุ่ม Prev/Next สำหรับเปลี่ยนเดือน
        - สถานะ: แดง = ติดจอง, เขียว/น้ำเงิน = มีจองแต่ยังไม่โอน, เหลือง = วันหยุด
        - ⚠️ ต้องกรองวันของเดือนอื่นที่แสดงในปฏิทินออก
        """
        print("🔄 กำลังโหลดหน้า Pattaya Party Pool Villa...")
        
        results = []
        
        # ดึงรหัสบ้านจาก URL
        match = re.search(r'/v/(\d+)', url)
        if match:
            villa_id = match.group(1)
            dv_code = f"DV-{villa_id}"
        else:
            villa_id = "Unknown"
            dv_code = "Unknown"
        
        self.driver.get(url)
        time.sleep(5)  # รอ JavaScript โหลด
        
        try:
            wait = WebDriverWait(self.driver, 15)
            
            # ดึงชื่อบ้านจาก header หรือ title
            try:
                # หารหัสที่พัก
                code_el = self.driver.find_element(By.XPATH, "//*[contains(text(),'รหัสที่พัก')]")
                house_info = code_el.text
                # หาชื่อจาก title
                title = self.driver.title
                house_name = title.split('|')[0].strip() if '|' in title else title
            except:
                house_name = f"Villa {villa_id}"
            
            print(f"  🏠 บ้าน: {house_name} ({dv_code})")
            
            # กดปุ่ม "📅 วันนี้" เพื่อกลับไปเดือนปัจจุบันก่อน
            try:
                today_btn = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//button[contains(text(),'วันนี้') or contains(@title,'กลับไปเดือนปัจจุบัน')]")
                    )
                )
                today_btn.click()
                time.sleep(2)
            except:
                pass  # ถ้าไม่มีปุ่มก็ข้ามไป
            
            # ดึงปฏิทินหลายเดือน
            start_date = datetime.now()
            
            for i in range(MONTH_TO_SCRAPE):
                try:
                    if i > 0:
                        # กดปุ่ม Next เพื่อไปเดือนถัดไป
                        try:
                            next_btn = wait.until(
                                EC.element_to_be_clickable(
                                    (By.XPATH, "//button[contains(text(),'Next') or contains(text(),'►') or contains(text(),'>')]")
                                )
                            )
                            next_btn.click()
                            time.sleep(2)  # รอปฏิทินโหลดใหม่
                        except Exception as e:
                            print(f"  ⚠️ ไม่สามารถกดปุ่ม Next: {e}")
                            break
                    
                    # คำนวณเดือน/ปีที่คาดหวัง
                    target_date = start_date + relativedelta(months=i)
                    expected_month = target_date.month
                    expected_year = target_date.year
                    
                    # อ่านชื่อเดือนจาก header ปฏิทิน
                    month_text = ""
                    try:
                        month_el = self.driver.find_element(
                            By.XPATH,
                            "//*[contains(text(),'มกราคม') or contains(text(),'กุมภาพันธ์') or contains(text(),'มีนาคม') or contains(text(),'เมษายน') or contains(text(),'พฤษภาคม') or contains(text(),'มิถุนายน') or contains(text(),'กรกฎาคม') or contains(text(),'สิงหาคม') or contains(text(),'กันยายน') or contains(text(),'ตุลาคม') or contains(text(),'พฤศจิกายน') or contains(text(),'ธันวาคม')]"
                        )
                        month_text = month_el.text.strip()
                        # ดึงเฉพาะส่วนที่มีเดือนและปี
                        for line in month_text.split('\n'):
                            if any(m in line for m in ['มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน', 'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม']):
                                if '256' in line or '257' in line:
                                    month_text = line.strip()
                                    break
                    except:
                        month_text = target_date.strftime("%Y-%m")
                    
                    # หาจำนวนวันในเดือนปัจจุบัน
                    import calendar
                    days_in_month = calendar.monthrange(expected_year, expected_month)[1]
                    
                    # ดึงวันที่ติดจอง - หา div ที่มีสีแดงและเป็นวันของเดือนปัจจุบัน
                    booked_days = []
                    
                    # เว็บนี้ใช้ div แทน table!
                    # วันที่ติดจอง = มี class bg-red-500 และ text-white
                    # วันของเดือนอื่น = มี class text-gray-400
                    
                    # หา div ที่เป็นวันในปฏิทิน (มี class grid และ grid-cols-7)
                    # มี 2 อัน: อันแรก = header (อา.จ.อ.พ.พฤ.ศ.ส.), อันที่สอง = ตัวเลขวัน
                    try:
                        calendar_grids = self.driver.find_elements(
                            By.CSS_SELECTOR,
                            "div.grid.grid-cols-7"
                        )
                        # ใช้ grid ที่สอง (index 1) ถ้ามี
                        calendar_grid = calendar_grids[1] if len(calendar_grids) > 1 else calendar_grids[0]
                        
                        # หา cell ทั้งหมดที่เป็นวัน
                        day_cells = calendar_grid.find_elements(
                            By.CSS_SELECTOR,
                            "div.aspect-square"
                        )
                    except:
                        # fallback - หา div ที่มี bg-red-500
                        day_cells = self.driver.find_elements(
                            By.CSS_SELECTOR,
                            "div.aspect-square"
                        )
                    
                    for cell in day_cells:
                        try:
                            class_attr = cell.get_attribute("class") or ""
                            cell_text = cell.text.strip()
                            
                            # ข้ามถ้าเป็นวันของเดือนอื่น (มี text-gray-400)
                            if "text-gray-400" in class_attr or "text-gray" in class_attr:
                                continue
                            
                            # ตรวจสอบว่าเป็นวันติดจอง (สีแดง: bg-red-500)
                            is_booked = "bg-red" in class_attr
                            
                            if is_booked and cell_text:
                                # ดึงเลขวัน
                                numbers = re.findall(r'\d+', cell_text)
                                if numbers:
                                    day = int(numbers[0])
                                    if 1 <= day <= days_in_month:
                                        booked_days.append(day)
                        except:
                            pass
                    
                    # ลบรายการซ้ำ
                    booked_days = sorted(set(booked_days))
                    
                    # เพิ่มลง results
                    for day in booked_days:
                        results.append({
                            "เว็บไซต์": "Pattaya Party Pool Villa",
                            "ชื่อบ้าน": house_name,
                            "รหัส": dv_code,
                            "เดือน": month_text,
                            "วันที่": day,
                            "สถานะ": "ติดจอง"
                        })
                    
                    if booked_days:
                        days_str = ', '.join(map(str, booked_days))
                        print(f"  📅 {month_text}: {len(booked_days)} วัน → [{days_str}]")
                    else:
                        print(f"  📅 {month_text}: ว่าง ✓")
                        
                except Exception as e:
                    print(f"  ⛔ Error เดือนที่ {i+1}: {e}")
            
            # ถ้าไม่พบข้อมูล แสดง debug info
            if not results:
                print("\n  ⚠️ ไม่พบข้อมูลการจอง - กำลัง debug...")
                self._debug_calendar_structure()
                
        except Exception as e:
            print(f"  ⛔ Error: {e}")
        
        return results
    
    
    def _debug_calendar_structure(self):
        """แสดง debug info สำหรับวิเคราะห์โครงสร้างปฏิทิน"""
        print("\n  📋 Debug: กำลังวิเคราะห์โครงสร้าง HTML...")
        
        # หา elements ที่อาจเป็นปฏิทิน
        tables = self.driver.find_elements(By.TAG_NAME, "table")
        print(f"  - พบ table: {len(tables)} อัน")
        
        # หา td ทั้งหมด
        tds = self.driver.find_elements(By.TAG_NAME, "td")
        print(f"  - พบ td: {len(tds)} อัน")
        
        # หา class ที่มี bg-
        bg_elements = self.driver.find_elements(By.XPATH, "//*[contains(@class,'bg-')]")
        classes = set()
        for el in bg_elements[:50]:  # จำกัด 50 อัน
            class_attr = el.get_attribute("class")
            if class_attr:
                for c in class_attr.split():
                    if 'bg-' in c:
                        classes.add(c)
        
        if classes:
            print(f"  - พบ background classes: {list(classes)[:10]}")
        
        # Save HTML สำหรับ debug
        try:
            with open("debug_calendar.html", "w", encoding="utf-8") as f:
                f.write(self.driver.page_source)
            print("  💾 บันทึก HTML ไว้ที่ debug_calendar.html")
        except:
            pass


def main():
    print("=" * 60)
    print("🏠 Pool Villa Calendar Scraper")
    print("📅 รองรับ 3 เว็บไซต์:")
    print("   1. devillegroups.com")
    print("   2. poolvillacity.co.th")
    print("   3. pattayapartypoolvilla.com")
    print("=" * 60)
    
    # ตั้งค่า Chrome
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    
    scraper = CalendarScraper(driver)
    all_results = []

    # โหลด URL จากไฟล์ webpath หากมี มิฉะนั้นใช้ URLS (fallback)
    urls_from_file = load_urls_from_webpath()
    urls_to_scrape = urls_from_file if urls_from_file else URLS

    print("\nURLs ที่จะดึง:")
    for i, u in enumerate(urls_to_scrape, 1):
        print(f"  {i}. {u}")

    # วน scrape แต่ละ URL
    for url in urls_to_scrape:
        try:
            results = scraper.scrape(url)
            all_results.extend(results)
        except Exception as e:
            print(f"❌ Error scraping {url}: {e}")

    driver.quit()

    # กรองวันที่ก่อนวันปัจจุบันออก
    if all_results:
        original_count = len(all_results)
        all_results = scraper.filter_past_dates(all_results)
        filtered_count = original_count - len(all_results)
        if filtered_count > 0:
            print(f"\n🗑️ กรองวันที่ผ่านมาแล้วออก: {filtered_count} รายการ")

    # Export ผลลัพธ์
    if all_results:
        df = pd.DataFrame(all_results)
        
        # บันทึก CSV
        df.to_csv("booking_result.csv", index=False, encoding="utf-8-sig")
        
        # บันทึก Excel พร้อมจัดรูปแบบ
        try:
            # ลองใช้ xlsxwriter ก่อน ถ้าไม่มีค่อยใช้ openpyxl
            try:
                import xlsxwriter
                engine = "xlsxwriter"
            except ImportError:
                try:
                    import openpyxl
                    engine = "openpyxl"
                except ImportError:
                    engine = None
                    print("⚠️ ไม่พบ openpyxl/xlsxwriter - ข้ามการสร้าง Excel")
            
            if engine:
                xlsx_file = "booking_result.xlsx"
                with pd.ExcelWriter(xlsx_file, engine=engine) as writer:
                    sheet_name = "Bookings"
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    
                    # จัดรูปแบบ
                    try:
                        worksheet = writer.sheets[sheet_name]
                        
                        # Freeze หัวตาราง
                        if engine == "xlsxwriter":
                            worksheet.freeze_panes(1, 0)
                        else:  # openpyxl
                            worksheet.freeze_panes = "A2"
                        
                        # ตั้งความกว้างคอลัมน์
                        col_widths = {
                            "เว็บไซต์": 22,
                            "ชื่อบ้าน": 30,
                            "รหัส": 14,
                            "เดือน": 18,
                            "วันที่": 10,
                            "สถานะ": 12,
                        }
                        
                        for idx, col in enumerate(df.columns):
                            width = col_widths.get(col, max(10, min(40, len(col) + 2)))
                            if engine == "xlsxwriter":
                                worksheet.set_column(idx, idx, width)
                            else:
                                from openpyxl.utils import get_column_letter
                                worksheet.column_dimensions[get_column_letter(idx + 1)].width = width
                        
                        # AutoFilter
                        if engine == "xlsxwriter":
                            worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)
                        else:
                            from openpyxl.utils import get_column_letter
                            last_col = get_column_letter(len(df.columns))
                            last_row = len(df) + 1
                            worksheet.auto_filter.ref = f"A1:{last_col}{last_row}"
                    except Exception as e:
                        print(f"⚠️ จัดรูปแบบ Excel ไม่สำเร็จ: {e}")
                
                print(f"\n{'='*60}")
                print(f"✅ เสร็จสิ้น!")
                print(f"📊 รวมข้อมูล: {len(all_results)} รายการ (หลังกรอง)")
                print(f"💾 บันทึกไฟล์:")
                print(f"   📄 booking_result.csv")
                print(f"   📊 booking_result.xlsx")
                print(f"{'='*60}")
            else:
                print(f"\n{'='*60}")
                print(f"✅ เสร็จสิ้น!")
                print(f"📊 รวมข้อมูล: {len(all_results)} รายการ (หลังกรอง)")
                print(f"💾 บันทึกไฟล์: booking_result.csv")
                print(f"{'='*60}")
        except Exception as e:
            print(f"⚠️ สร้าง Excel ไม่สำเร็จ: {e}")
            print(f"\n{'='*60}")
            print(f"✅ เสร็จสิ้น!")
            print(f"📊 รวมข้อมูล: {len(all_results)} รายการ (หลังกรอง)")
            print(f"💾 บันทึกไฟล์: booking_result.csv")
            print(f"{'='*60}")
        
        print(df.head(20))
    else:
        print(f"\n{'='*60}")
        print("⚠️ ไม่พบข้อมูลการจอง")
        print("💡 อาจต้องปรับ CSS selector ให้ตรงกับโครงสร้าง HTML ของเว็บ")
        print(f"{'='*60}")


if __name__ == "__main__":
    main()
